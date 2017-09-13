from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import PatternFill, Color
wb = Workbook()
ws = wb.active
ws['B2'] = 'Hello'

ws.title = 'Text'

cal = ws['B2']
cal.font = Font(name='맑은고딕', size=15, bold=True)
cal.alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('B2:E2')
title = ws['B2:E2']
box = Border(left=Side(border_style="thin", color="FF000000"),
             right=Side(border_style="thin", color="FF000000"),
             top=Side(border_style="thin", color="FF000000"),
             bottom=Side(border_style="thin", color="FF000000"))
title.border = box

title.fill = PatternFill(patternType='solid', fgColor=Color('FFC000'))

wb.save('text.xlsx')